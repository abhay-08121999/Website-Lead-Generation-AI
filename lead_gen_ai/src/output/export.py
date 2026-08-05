"""
export.py
---------
Exports the final, scored lead list to a polished Excel file:
sorted by priority score, color-coded by lead category, with
auto-sized columns and a summary sheet.
"""

import os
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CATEGORY_COLORS = {
    "NO_WEBSITE": "FFC7CE",    # red-ish -> hottest lead
    "POOR_WEBSITE": "FFEB9C",  # yellow -> warm lead
}

COLUMNS = [
    ("business_name", "Business Name"),
    ("category", "Category"),
    ("city", "City"),
    ("phone", "Phone"),
    ("address", "Address"),
    ("rating", "Rating"),
    ("review_count", "Reviews"),
    ("website", "Website URL"),
    ("website_status", "Website Status"),
    ("lead_category", "Lead Type"),
    ("lead_score", "Priority Score"),
    ("lead_reason", "Why This Is A Lead"),
    ("source", "Data Source"),
]


def export_leads_to_excel(leads: List[Dict], output_path: str) -> str:
    """
    Writes leads to an .xlsx file with:
      - Sheet 1: "Leads" — full sorted, color-coded list
      - Sheet 2: "Summary" — counts by city / category / lead type
    Returns the output path.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Sort: highest priority score first
    sorted_leads = sorted(leads, key=lambda x: x.get("lead_score", 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, (_, header_label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, lead in enumerate(sorted_leads, start=2):
        lead_cat = lead.get("lead_category", "")
        fill_color = CATEGORY_COLORS.get(lead_cat)
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            value = lead.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Auto-size columns (approx, based on content length)
    for col_idx, (key, header_label) in enumerate(COLUMNS, start=1):
        max_len = len(header_label)
        for lead in sorted_leads:
            val = str(lead.get(key, ""))
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Metric", "Value"])
    summary_ws["A1"].font = header_font
    summary_ws["B1"].font = header_font
    summary_ws["A1"].fill = header_fill
    summary_ws["B1"].fill = header_fill

    total = len(sorted_leads)
    no_website_count = sum(1 for l in sorted_leads if l.get("lead_category") == "NO_WEBSITE")
    poor_website_count = sum(1 for l in sorted_leads if l.get("lead_category") == "POOR_WEBSITE")

    summary_rows = [
        ("Total Qualified Leads", total),
        ("No Website (Hottest)", no_website_count),
        ("Poor/Underperforming Website (Warm)", poor_website_count),
        ("", ""),
        ("Leads by City", ""),
    ]

    cities = {}
    for l in sorted_leads:
        c = l.get("city", "Unknown")
        cities[c] = cities.get(c, 0) + 1
    for city, count in sorted(cities.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {city}", count))

    summary_rows.append(("", ""))
    summary_rows.append(("Leads by Category", ""))
    cats = {}
    for l in sorted_leads:
        c = l.get("category", "Unknown")
        cats[c] = cats.get(c, 0) + 1
    for cat, count in sorted(cats.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {cat}", count))

    for row in summary_rows:
        summary_ws.append(row)

    summary_ws.column_dimensions["A"].width = 40
    summary_ws.column_dimensions["B"].width = 15

    wb.save(output_path)
    return output_path
